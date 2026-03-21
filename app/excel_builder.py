import os
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side,
)
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter

from app.config import OUTPUT_DIR

EXCEL_THEMES = {
    "blue": {
        "primary": "0D47A1",
        "accent": "29B6F6",
        "header_bg": "0D47A1",
        "header_fg": "FFFFFF",
        "alt_row": "E3F2FD",
        "border": "B0BEC5",
        "kpi_bg": "E8F0FE",
        "chart_colors": ["29B6F6", "0D47A1", "0277BD", "4FC3F7", "039BE5"],
    },
    "green": {
        "primary": "1B5E20",
        "accent": "66BB6A",
        "header_bg": "1B5E20",
        "header_fg": "FFFFFF",
        "alt_row": "E8F5E9",
        "border": "A5D6A7",
        "kpi_bg": "E8F5E9",
        "chart_colors": ["2E7D32", "43A047", "66BB6A", "1B5E20", "81C784"],
    },
    "orange": {
        "primary": "E65100",
        "accent": "FF8F00",
        "header_bg": "E65100",
        "header_fg": "FFFFFF",
        "alt_row": "FFF3E0",
        "border": "FFCC80",
        "kpi_bg": "FFF8E1",
        "chart_colors": ["FF8F00", "E65100", "BF360C", "FFB74D", "FF6D00"],
    },
    "purple": {
        "primary": "4A148C",
        "accent": "CE93D8",
        "header_bg": "4A148C",
        "header_fg": "FFFFFF",
        "alt_row": "F3E5F5",
        "border": "CE93D8",
        "kpi_bg": "F3E5F5",
        "chart_colors": ["7B1FA2", "CE93D8", "4A148C", "E1BEE7", "AB47BC"],
    },
    "gray": {
        "primary": "37474F",
        "accent": "78909C",
        "header_bg": "37474F",
        "header_fg": "FFFFFF",
        "alt_row": "ECEFF1",
        "border": "B0BEC5",
        "kpi_bg": "ECEFF1",
        "chart_colors": ["546E7A", "78909C", "37474F", "90A4AE", "607D8B"],
    },
}

_FONT = "Yu Gothic UI"


def _theme(name: str) -> dict:
    return EXCEL_THEMES.get(name, EXCEL_THEMES["blue"])


def _border(color: str) -> Border:
    side = Side(style="thin", color=color)
    return Border(left=side, right=side, top=side, bottom=side)


def build_excel(data: dict, filename: str) -> str:
    wb = Workbook()
    t = _theme(data.get("theme", "blue"))
    sheets = data.get("sheets", [])

    for idx, sheet_data in enumerate(sheets):
        name = str(sheet_data.get("name", f"Sheet{idx + 1}"))[:31]
        if idx == 0:
            ws = wb.active
            ws.title = name
        else:
            ws = wb.create_sheet(title=name)
        _build_sheet(ws, sheet_data, t)

    if not sheets:
        ws = wb.active
        ws.title = "レポート"
        ws["A1"] = data.get("title", "レポート")
        ws["A1"].font = Font(name=_FONT, size=16, bold=True)

    safe = "".join(
        c for c in filename
        if c.isalnum() or c in (" ", "_", "-", ".", "ー")
        or ("\u3000" <= c <= "\u9fff") or ("\uac00" <= c <= "\ud7af")
    ).strip() or "report"
    path = os.path.join(OUTPUT_DIR, f"{safe}.xlsx")
    wb.save(path)
    return path


def _build_sheet(ws, sheet_data: dict, t: dict):
    sections = sheet_data.get("sections", [])
    row = 1
    for sec in sections:
        kind = sec.get("type", "text")
        handler = _SECTION_HANDLERS.get(kind, _sec_text)
        row = handler(ws, sec, t, row)
        row += 1


# ---------------------------------------------------------------------------
# Section handlers
# ---------------------------------------------------------------------------

def _sec_title(ws, sec: dict, t: dict, row: int) -> int:
    title = sec.get("text", "")
    subtitle = sec.get("subtitle", "")

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws.cell(row=row, column=1, value=title)
    c.font = Font(name=_FONT, size=20, bold=True, color=t["primary"])
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 40
    row += 1

    if subtitle:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        c = ws.cell(row=row, column=1, value=subtitle)
        c.font = Font(name=_FONT, size=12, color="666666")
        row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws.cell(row=row, column=1,
                value=f"作成日: {datetime.now().strftime('%Y年%m月%d日')}")
    c.font = Font(name=_FONT, size=10, color="999999")
    row += 1

    bar_fill = PatternFill(start_color=t["accent"], end_color=t["accent"],
                           fill_type="solid")
    for col in range(1, 9):
        ws.cell(row=row, column=col).fill = bar_fill
    ws.row_dimensions[row].height = 4
    return row + 1


def _sec_heading(ws, sec: dict, t: dict, row: int) -> int:
    text = sec.get("text", "")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    c = ws.cell(row=row, column=1, value=f"■ {text}")
    c.font = Font(name=_FONT, size=14, bold=True, color=t["primary"])
    c.alignment = Alignment(vertical="center")
    ws.row_dimensions[row].height = 30
    row += 1

    bar_fill = PatternFill(start_color=t["accent"], end_color=t["accent"],
                           fill_type="solid")
    for col in range(1, 4):
        ws.cell(row=row, column=col).fill = bar_fill
    ws.row_dimensions[row].height = 3
    return row + 1


def _sec_text(ws, sec: dict, t: dict, row: int) -> int:
    content = sec.get("content", "")
    for line in content.split("\n"):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        c = ws.cell(row=row, column=1, value=line)
        c.font = Font(name=_FONT, size=11)
        c.alignment = Alignment(wrap_text=True, vertical="top")
        row += 1
    return row


def _sec_table(ws, sec: dict, t: dict, row: int) -> int:
    headers = sec.get("headers", [])
    rows_data = sec.get("rows", [])
    table_title = sec.get("title", "")
    num_cols = len(headers)

    if table_title:
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=max(num_cols, 1))
        c = ws.cell(row=row, column=1, value=table_title)
        c.font = Font(name=_FONT, size=12, bold=True, color=t["primary"])
        row += 1

    if not headers:
        return row

    bdr = _border(t["border"])
    h_fill = PatternFill(start_color=t["header_bg"],
                         end_color=t["header_bg"], fill_type="solid")
    h_font = Font(name=_FONT, size=11, bold=True, color=t["header_fg"])

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.font = h_font
        c.fill = h_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = bdr
        ws.column_dimensions[get_column_letter(ci)].width = max(
            14, min(30, len(str(h)) * 2 + 4)
        )
    ws.row_dimensions[row].height = 28
    row += 1

    alt_fill = PatternFill(start_color=t["alt_row"],
                           end_color=t["alt_row"], fill_type="solid")
    body_font = Font(name=_FONT, size=11)

    for ri, rd in enumerate(rows_data):
        for ci, val in enumerate(rd, 1):
            if ci > num_cols:
                break
            c = ws.cell(row=row, column=ci, value=val)
            c.font = body_font
            c.border = bdr
            c.alignment = Alignment(vertical="center")
            if ri % 2 == 1:
                c.fill = alt_fill
            if isinstance(val, (int, float)):
                c.alignment = Alignment(horizontal="right", vertical="center")
                c.number_format = "#,##0" if isinstance(val, int) else "#,##0.00"
        row += 1
    return row


def _sec_chart(ws, sec: dict, t: dict, row: int) -> int:
    chart_type = sec.get("chart_type", "bar")
    chart_title = sec.get("title", "")
    categories = sec.get("categories", [])
    series_list = sec.get("series", [])

    if not categories or not series_list:
        return row

    data_start = row
    ws.cell(row=row, column=1, value="")
    for si, s in enumerate(series_list):
        ws.cell(row=row, column=si + 2, value=s.get("name", f"S{si + 1}"))
    row += 1

    for ci, cat in enumerate(categories):
        ws.cell(row=row, column=1, value=cat)
        for si, s in enumerate(series_list):
            vals = s.get("values", [])
            if ci < len(vals):
                ws.cell(row=row, column=si + 2, value=vals[ci])
        row += 1

    data_end = row - 1
    num_s = len(series_list)

    if chart_type == "pie":
        ch = PieChart()
    elif chart_type == "line":
        ch = LineChart()
        ch.y_axis.numFmt = "#,##0"
    else:
        ch = BarChart()
        ch.y_axis.numFmt = "#,##0"

    ch.title = chart_title
    ch.style = 10
    ch.width = 18
    ch.height = 12

    data_ref = Reference(ws, min_col=2, max_col=1 + num_s,
                         min_row=data_start, max_row=data_end)
    cats_ref = Reference(ws, min_col=1,
                         min_row=data_start + 1, max_row=data_end)
    ch.add_data(data_ref, titles_from_data=True)
    ch.set_categories(cats_ref)

    colors = t.get("chart_colors", ["4472C4"])
    for i, s in enumerate(ch.series):
        s.graphicalProperties.solidFill = colors[i % len(colors)]

    ws.add_chart(ch, f"A{row}")
    return row + 16


def _sec_kpi(ws, sec: dict, t: dict, row: int) -> int:
    items = sec.get("items", [])
    kpi_title = sec.get("title", "")
    n = len(items) if items else 0

    if kpi_title:
        ws.merge_cells(start_row=row, start_column=1,
                       end_row=row, end_column=max(n * 2, 2))
        c = ws.cell(row=row, column=1, value=kpi_title)
        c.font = Font(name=_FONT, size=12, bold=True, color=t["primary"])
        row += 1

    if not items:
        return row

    bdr = _border(t["border"])
    label_fill = PatternFill(start_color=t["kpi_bg"],
                             end_color=t["kpi_bg"], fill_type="solid")
    val_fill = PatternFill(start_color=t["primary"],
                           end_color=t["primary"], fill_type="solid")

    for i, item in enumerate(items):
        col = i * 2 + 1
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + 1)
        c = ws.cell(row=row, column=col, value=item.get("label", ""))
        c.font = Font(name=_FONT, size=10, color="666666")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = label_fill
        c.border = bdr
        ws.cell(row=row, column=col + 1).fill = label_fill
        ws.cell(row=row, column=col + 1).border = bdr
    ws.row_dimensions[row].height = 24
    row += 1

    for i, item in enumerate(items):
        col = i * 2 + 1
        ws.merge_cells(start_row=row, start_column=col,
                       end_row=row, end_column=col + 1)
        c = ws.cell(row=row, column=col, value=item.get("value", ""))
        c.font = Font(name=_FONT, size=18, bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.fill = val_fill
        c.border = bdr
        ws.cell(row=row, column=col + 1).fill = val_fill
        ws.cell(row=row, column=col + 1).border = bdr
    ws.row_dimensions[row].height = 45
    return row + 1


_SECTION_HANDLERS = {
    "title": _sec_title,
    "heading": _sec_heading,
    "text": _sec_text,
    "table": _sec_table,
    "chart": _sec_chart,
    "kpi": _sec_kpi,
}
